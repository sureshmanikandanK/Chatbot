# # preprocessing_file

# # for pdf i will use the pyupdf4ll to make an pdf to markdown and chunk the data 
# # import PyPDF2
# from langchain_community.document_loaders import PyPDFLoader
# import pymupdf4llm
# from langchain.text_splitter import MarkdownTextSplitter
# import os   
# from langchain.text_splitter import RecursiveCharacterTextSplitter
# from langchain_google_genai import GoogleGenerativeAIEmbeddings
# from langchain_community.vectorstores import FAISS
# from langchain_community.document_loaders import PyMuPDFLoader
# from langchain_community.document_loaders.csv_loader import CSVLoader
# partent_director_index = "test_files/uploads/index"



# def csv_documet_load(file_path):
#     loader = CSVLoader(file_path)
#     data = loader.load_and_split() 
#     return data

# # this functuon load the document and make an pdf as an markdown file and chunked the data by using 500 characters
# def load_pdf_doc_split(file_path):
#     loader = PyMuPDFLoader(file_path)
#     data = loader.load()
#     # mark_down = pymupdf4llm.to_markdown(file_path)
#     # split_markdown = MarkdownTextSplitter(chunk_size=500, chunk_overlap=50)
#     # split_markdown = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
    
#     # chunked = split_markdown.split_documents(data)
#     return data

# def embadding_and_store(chunked, file_name, bot_name, user_id, file_id):
#     # Use embeddings with user-specific metadata
#     model_name = "models/embedding-001"
#     api_key = "AIzaSyA3-0m7MQxNzWZJzVyyOCmNlU_Z_q-SWiw"
#     embeddings = GoogleGenerativeAIEmbeddings(model=model_name, google_api_key=api_key)  # Create embeddings
 
#     # Add user-specific metadata to each chunk
#     for doc in chunked:
#         doc.metadata['user_id'] = user_id  # Associate user ID
#         doc.metadata['file_name'] = file_name  # Optionally associate file name
#         doc.metadata['file_id'] = file_id
#     # Create a single shared FAISS index
#     db = FAISS.from_documents(chunked, embeddings)

#     # Save to a shared directory
#     path = f"./uploads/shared_index"
#     if not os.path.exists(path):
#         os.makedirs(path) 

#     db.save_local(path)
#     return path

# _________original is just above _________

# from langchain_community.document_loaders import PyPDFLoader, CSVLoader
# from langchain_community.document_loaders.csv_loader import CSVLoader
# from langchain_community.document_loaders.docx_loader import DOCXLoader  # New loader for DOCX
# import pandas as pd  # For Excel
# import os
# from langchain.text_splitter import MarkdownTextSplitter, RecursiveCharacterTextSplitter
# from langchain_google_genai import GoogleGenerativeAIEmbeddings
# from langchain_community.vectorstores import FAISS
# from langchain_community.document_loaders import PyMuPDFLoader

# partent_director_index = "test_files/uploads/index"

# # Function to load and process CSV files
# def csv_document_load(file_path):
#     loader = CSVLoader(file_path)
#     data = loader.load_and_split() 
#     return data

# # Function to load and process Excel files
# def excel_document_load(file_path):
#     # Load the Excel file with pandas
#     df = pd.read_excel(file_path, sheet_name=None)  # Load all sheets
#     documents = []
#     for sheet_name, sheet_data in df.items():
#         # Convert each sheet to text and chunk it
#         documents.append(sheet_data.to_string())
#     return documents

# # Function to load and process DOCX files
# def docx_document_load(file_path):
#     loader = DOCXLoader(file_path)
#     data = loader.load()
#     return data

# # Function to load and process TXT files
# def txt_document_load(file_path):
#     with open(file_path, 'r') as file:
#         text = file.read()
#     return [text]  # Return as a list of one document

# # Function to load and process PDF files
# def load_pdf_doc_split(file_path):
#     loader = PyMuPDFLoader(file_path)
#     data = loader.load()
#     return data

# # Generic document loader based on file type
# def load_document(file_path):
#     file_extension = file_path.split('.')[-1].lower()

#     if file_extension == 'csv':
#         return csv_document_load(file_path)
#     elif file_extension == 'xlsx' or file_extension == 'xls':
#         return excel_document_load(file_path)
#     elif file_extension == 'docx':
#         return docx_document_load(file_path)
#     elif file_extension == 'txt':
#         return txt_document_load(file_path)
#     elif file_extension == 'pdf':
#         return load_pdf_doc_split(file_path)
#     else:
#         raise ValueError(f"Unsupported file type: {file_extension}")

# # Function to create embeddings and store the documents
# def embadding_and_store(chunked, file_name, bot_name, user_id, file_id):
#     model_name = "models/embedding-001"
#     api_key = "AIzaSyA3-0m7MQxNzWZJzVyyOCmNlU_Z_q-SWiw"
#     embeddings = GoogleGenerativeAIEmbeddings(model=model_name, google_api_key=api_key)

#     # Add user-specific metadata to each chunk
#     for doc in chunked:
#         doc.metadata['user_id'] = user_id
#         doc.metadata['file_name'] = file_name
#         doc.metadata['file_id'] = file_id
    
#     # Create a FAISS index and store it
#     db = FAISS.from_documents(chunked, embeddings)
    
#     # Save to a shared directory
#     path = f"./uploads/shared_index"
#     if not os.path.exists(path):
#         os.makedirs(path) 

#     db.save_local(path)
#     return path




# preprocessing_file.py
import os
from langchain_community.document_loaders import CSVLoader, PyMuPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain_google_genai import GoogleGenerativeAIEmbeddings

# Additional libraries for DOCX, TXT, and Excel files
import openpyxl
import docx

partent_director_index = "test_files/uploads/index"

def get_file_loader(file_path):
    # Determine the file extension to select the correct loader
    file_extension = file_path.split('.')[-1].lower()

    if file_extension == 'csv':
        return CSVLoader(file_path)
    elif file_extension == 'pdf':
        return PyMuPDFLoader(file_path)
    elif file_extension == 'xlsx' or file_extension == 'xls':
        return load_excel_document(file_path)
    elif file_extension == 'docx':
        return load_docx_document(file_path)
    elif file_extension == 'txt':
        return load_txt_document(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_extension}")

def csv_documet_load(file_path):
    try:
        loader = get_file_loader(file_path)
        if isinstance(loader, CSVLoader):
            data = loader.load_and_split()
        else:
            # If it's a PDF or other document type, we need to handle accordingly
            data = loader.load()  # Get the document contents
            # Optionally, split the documents if needed
            splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
            data = splitter.split_documents(data)  # Split into chunks if needed
        return data
    except Exception as e:
        print(f"Error loading file {file_path}: {str(e)}")
        raise RuntimeError(f"Error loading file {file_path}: {str(e)}")

# Function to load PDF documents
def load_pdf_doc_split(file_path):
    loader = PyMuPDFLoader(file_path)
    data = loader.load()
    # Split the loaded PDF content into chunks
    splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
    data = splitter.split_documents(data)
    return data

# Function to load Excel documents
def load_excel_document(file_path):
    workbook = openpyxl.load_workbook(file_path)
    data = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            data.append(" ".join([str(cell) for cell in row if cell]))
    return data

# Function to load DOCX documents
def load_docx_document(file_path):
    doc = docx.Document(file_path)
    data = []
    for para in doc.paragraphs:
        data.append(para.text)
    return data

# Function to load TXT documents
def load_txt_document(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = file.readlines()
    return data

def embadding_and_store(chunked, file_name, bot_name, user_id, file_id):
    # Use embeddings with user-specific metadata
    model_name = "models/embedding-001"
    api_key = "AIzaSyA3-0m7MQxNzWZJzVyyOCmNlU_Z_q-SWiw"
    embeddings = GoogleGenerativeAIEmbeddings(model=model_name, google_api_key=api_key)  # Create embeddings
 
    # Add user-specific metadata to each chunk
    for doc in chunked:
        doc.metadata['user_id'] = user_id  # Associate user ID
        doc.metadata['file_name'] = file_name  # Optionally associate file name
        doc.metadata['file_id'] = file_id

    # Create a single shared FAISS index
    db = FAISS.from_documents(chunked, embeddings)

    # Save to a shared directory
    path = f"./uploads/shared_index"
    if not os.path.exists(path):
        os.makedirs(path)

    db.save_local(path)
    return path
